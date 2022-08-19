from openpyxl import load_workbook
from openpyxl.chart import BarChart , Reference

wb = load_workbook(
    "Free_Code_Camp Full Automation\Project 3 Automate Excel Report\sources\pivot_table.xlsx")
sheet = wb['Report']

# first to get used columns and rows in sheet
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(min_column, max_column, min_row, max_row)

barchart = BarChart()

data = Reference(sheet,min_column+1,max_column,min_row+1,max_row) # we use +1 poeration to getrid of the headers
categories = Reference(sheet, min_column, min_column, min_row, min_row) # now we want to get headers

barchart.add_data(data,titles_from_data = True)
barchart.set_categories(categories)

sheet.add_chart(barchart,"B12")

barchart.title = "Sales by Product Line"
barchart.style = 5

wb.save("Free_Code_Camp Full Automation\Project 3 Automate Excel Report\sources\Barchart.xlsx")


